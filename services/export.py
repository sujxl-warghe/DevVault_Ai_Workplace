"""Excel/CSV export — Title, Details, Repo Link, Live Link, matching the
current minimal project schema (spec section 34)."""
from __future__ import annotations

import io
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

COLUMNS = ["Title", "Details", "Repo Link", "Live Link"]
COLUMN_WIDTHS = [32, 60, 36, 36]
FIELD_KEYS = ["title", "description", "github_url", "demo_url"]


def _today_stamp() -> str:
    return date.today().isoformat()


def build_filename(extension: str) -> str:
    return f"DevVault_Projects_{_today_stamp()}.{extension}"


def export_projects_to_excel_bytes(projects: list[dict[str, Any]]) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    link_columns = {3, 4}  # Repo Link, Live Link (1-indexed)
    for project in projects:
        row = [project.get(key) or "" for key in FIELD_KEYS]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in link_columns:
            value = row[col_idx - 1]
            if not value:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.hyperlink = value
            cell.font = Font(color="1155CC", underline="single")

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), build_filename("xlsx")


def export_projects_to_csv_bytes(projects: list[dict[str, Any]]) -> tuple[bytes, str]:
    import csv

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(COLUMNS)
    for project in projects:
        writer.writerow([project.get(key) or "" for key in FIELD_KEYS])

    return buffer.getvalue().encode("utf-8"), build_filename("csv")
